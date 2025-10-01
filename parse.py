import streamlit as st
import os, json, re, requests, time, tempfile
from datetime import datetime, timedelta, date
from lxml import etree
from io import BytesIO
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def fetch_all_notices_to_json(cpv_codes, date_start, date_end, buyer_country, json_path):
    API = "https://api.ted.europa.eu/v3/notices/search"
    PAYLOAD_BASE = {
        "query": f"(publication-date >={date_start}<={date_end}) AND (buyer-country IN ({buyer_country})) "
                 f"AND (classification-cpv IN ({cpv_codes})) "
                 "AND (notice-type IN (pin-cfc-standard pin-cfc-social qu-sy cn-standard cn-social subco cn-desg))",
        "fields": ["publication-number","links"],
        "scope": "ACTIVE",
        "checkQuerySyntax": False,
        "paginationMode": "PAGE_NUMBER",
        "page": 1,
        "limit": 100
    }
    s = requests.Session()
    s.headers.update({"Accept":"application/json"})
    all_notices = []
    page = 1
    while True:
        body = dict(PAYLOAD_BASE); body["page"] = page
        r = s.post(API, json=body, timeout=60)
        r.raise_for_status()
        data = r.json()
        notices = data.get("results") or data.get("items") or data.get("notices") or []
        if not notices:
            break
        all_notices.extend(notices)
        total = data.get("total") or data.get("totalCount")
        if not notices or (total and page * PAYLOAD_BASE["limit"] >= total):
            break
        page += 1
        time.sleep(0.25)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"notices": all_notices}, f, ensure_ascii=False, indent=2)


def _get_links_block(notice: dict) -> dict:
    links = notice.get("links") or {}
    if isinstance(links, dict) and "links" in links and isinstance(links["links"], dict):
        links = links["links"]
    if isinstance(links, dict):
        return { (k.lower() if isinstance(k,str) else k): v for k,v in links.items() }
    return {}


def _extract_xml_urls_from_notice(notice: dict) -> list:
    block = _get_links_block(notice)
    xml_block = block.get("xml")
    urls = []
    if isinstance(xml_block, dict):
        for k, v in xml_block.items():
            if isinstance(k, str) and k.lower() == "mul" and v:
                urls.append(v)
        for k, v in xml_block.items():
            if isinstance(k, str) and k.lower() != "mul" and v:
                urls.append(v)
    elif isinstance(xml_block, str) and xml_block:
        urls.append(xml_block)
    return urls


def fetch_notice_xml(session: requests.Session, pubno: str, notice: dict) -> bytes:
    xml_headers = {"Accept": "application/xml", "User-Agent": "Mozilla/5.0"}
    for url in _extract_xml_urls_from_notice(notice):
        try:
            r = session.get(url, headers=xml_headers, timeout=60)
            if r.status_code == 200 and r.content.strip():
                return r.content
        except requests.RequestException:
            pass
    for lang in ("en","de","fr"):
        url = f"https://ted.europa.eu/{lang}/notice/{pubno}/xml"
        try:
            r = session.get(url, headers=xml_headers, timeout=60)
            if r.status_code == 200 and r.content.strip():
                return r.content
        except requests.RequestException:
            pass
    detail_url = f"https://ted.europa.eu/en/notice/-/detail/{pubno}"
    try:
        html = session.get(detail_url, headers={"User-Agent":"Mozilla/5.0"}, timeout=60).text
        m = re.search(r'https://ted\.europa\.eu/(?:en|de|fr)/notice/' + re.escape(pubno) + r'/xml', html)
        if m:
            r = session.get(m.group(0), headers=xml_headers, timeout=60)
            if r.status_code == 200 and r.content.strip():
                return r.content
    except requests.RequestException:
        pass
    raise RuntimeError(f"No XML found for {pubno}")


def _first_text(nodes):
    for n in nodes or []:
        t = (n.text or "").strip()
        if t:
            return t
    return ""


def _norm_date(d: str) -> str:
    if not d:
        return ""
    d = d.rstrip("Zz")
    return d.split("T")[0].split("+")[0]


def _clean_title(raw: str) -> str:
    if not raw: return ""
    return re.sub(r"^\s*\d{4}[-_]\d{5,}[\s_\-–:]+", "", raw.strip())


def _parse_iso_date(d: str):
    try:
        return datetime.strptime(d, "%Y-%m-%d")
    except Exception:
        return None


def _duration_to_days(val: str, unit: str | None):
    if not val:
        return None
    try:
        num = float(str(val).strip().replace(",", "."))
    except Exception:
        return None
    u = (unit or "").upper()
    if u in ("DAY","D","DAYS"):
        return int(round(num))
    if u in ("MON","M","MONTH","MONTHS"):
        return int(round(num * 30))
    if u in ("ANN","Y","YEAR","YEARS"):
        return int(round(num * 365))
    return None

def parse_xml_fields(xml_bytes: bytes) -> dict:
    parser = etree.XMLParser(recover=True, huge_tree=True)
    root = etree.parse(BytesIO(xml_bytes), parser)
    ns = {k: v for k, v in (root.getroot().nsmap or {}).items() if k}
    ns.setdefault("cbc","urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2")
    ns.setdefault("cac","urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2")
    ns.setdefault("efac","http://data.europa.eu/p27/eforms-ubl-extension-aggregate-components/1")
    ns.setdefault("efbc","http://data.europa.eu/p27/eforms-ubl-extension-basic-components/1")
    out = {}
    out["Beschaffer"] = _first_text(
        root.xpath(".//cac:ContractingParty//cac:PartyName/cbc:Name", namespaces=ns)
        or root.xpath(".//efac:Organizations//efac:Company/cac:PartyName/cbc:Name", namespaces=ns)
    )
    out["Projektbezeichnung"] = _clean_title(
        _first_text(root.xpath(".//cac:ProcurementProject/cbc:Name | .//cbc:Title | .//efbc:Title", namespaces=ns))
    )
    out["Ort/Region"] = _first_text(root.xpath("//cac:PostalAddress[1]/cbc:CityName", namespaces=ns))
    out["Vergabeplattform"] = _first_text(
        root.xpath(".//cbc:AccessToolsURI | .//cbc:WebsiteURI | .//cbc:URI | .//cbc:EndpointID", namespaces=ns)
    )
    pub_id = _first_text(root.xpath(".//efbc:NoticePublicationID[@schemeName='ojs-notice-id']", namespaces=ns))
    out["Ted-Link"] = f"https://ted.europa.eu/en/notice/-/detail/{pub_id}" if pub_id else ""
    start_nodes = root.xpath(
        ".//cac:ProcurementProject/cac:PlannedPeriod/cbc:StartDate "
        "| .//cac:ProcurementProjectLot//cac:ProcurementProject/cac:PlannedPeriod/cbc:StartDate",
        namespaces=ns
    )
    end_nodes = root.xpath(
        ".//cac:ProcurementProject/cac:PlannedPeriod/cbc:EndDate "
        "| .//cac:ProcurementProjectLot//cac:ProcurementProject/cac:PlannedPeriod/cbc:EndDate",
        namespaces=ns
    )
    start_norm = _norm_date(_first_text(start_nodes))
    end_norm = _norm_date(_first_text(end_nodes))
    if not start_norm and end_norm:
        dur_nodes = root.xpath(
            ".//cac:ProcurementProject/cac:PlannedPeriod/cbc:DurationMeasure "
            "| .//cac:ProcurementProjectLot//cac:ProcurementProject/cac:PlannedPeriod/cbc:DurationMeasure",
            namespaces=ns
        )
        dur_val, dur_unit = None, None
        for dn in dur_nodes:
            text_val = (dn.text or "").strip()
            unit = (dn.get("unitCode") or "").strip()
            if text_val:
                dur_val, dur_unit = text_val, unit
                break
        days = _duration_to_days(dur_val, dur_unit) if dur_val else None
        if days:
            end_dt = _parse_iso_date(end_norm)
            if end_dt:
                start_norm = (end_dt - timedelta(days=days)).strftime("%Y-%m-%d")
    out["Projektstart"] = start_norm
    out["Projektende"] = end_norm
    crit_nodes = root.xpath(
        ".//*[contains(local-name(),'SelectionCriteria') or contains(local-name(),'SelectionCriterion')]/cbc:Description",
        namespaces=ns
    )
    crit_text = " ".join((n.text or "").strip() for n in crit_nodes if (n.text or "").strip())
    crit_text = re.sub(r"\bslc-[a-z0-9\-]+\b", "", crit_text, flags=re.I).strip()
    out["Geforderte Unternehmensreferenzen"] = crit_text
    out["Geforderte Kriterien CVs"] = "CV" if re.search(
        r"\b(CV|Lebenslauf|Schlüsselpersonal|key staff|personaleinsatz)\b", crit_text, re.I
    ) else ""
    amount_nodes = root.xpath(
        ".//cbc:EstimatedOverallContractAmount | .//cbc:EstimatedOverallContractAmount/cbc:Value | .//efbc:EstimatedValue | .//cbc:PayableAmount",
        namespaces=ns
    )
    value_text = ""
    if amount_nodes:
        for node in amount_nodes:
            if node.text and node.text.strip():
                value_text = node.text.strip()
                parent = node.getparent()
                currency = node.get("currencyID") or (parent.get("currencyID") if parent is not None else None)
                if currency:
                    value_text += f" {currency}"
                break
    out["Projektvolumen"] = value_text or ""

    # --- Main fix for Frist Abgabedatum: try both cac: and efac: InterestExpressionReceptionPeriod ---
    tender_deadline_date = _norm_date(
        _first_text(root.xpath(".//cac:TenderSubmissionDeadlinePeriod/cbc:EndDate", namespaces=ns))
    )
    if not tender_deadline_date:
        tender_deadline_date = _norm_date(
            _first_text(root.xpath(".//cac:TenderingTerms/cbc:SubmissionDeadlineDate", namespaces=ns))
        )
    if not tender_deadline_date:
        tender_deadline_date = _norm_date(
            _first_text(root.xpath(".//cac:InterestExpressionReceptionPeriod/cbc:EndDate", namespaces=ns))
        )
    if not tender_deadline_date:
        tender_deadline_date = _norm_date(
            _first_text(root.xpath(".//efac:InterestExpressionReceptionPeriod/cbc:EndDate", namespaces=ns))
        )
    participation_deadline_date = _norm_date(
        _first_text(root.xpath(".//cac:ParticipationRequestReceptionPeriod/cbc:EndDate", namespaces=ns))
    )
    if not participation_deadline_date:
        participation_deadline_date = _norm_date(
            _first_text(root.xpath(".//efac:ParticipationRequestReceptionPeriod/cbc:EndDate", namespaces=ns))
        )
    out["Frist Abgabedatum"] = tender_deadline_date or participation_deadline_date

    # Veröffentlichung Datum
    pub_date = _first_text(root.xpath(".//efbc:PublicationDate", namespaces=ns))
    if not pub_date:
        pub_date = _first_text(root.xpath(".//cbc:PublicationDate", namespaces=ns))
    out["Veröffentlichung Datum"] = _norm_date(pub_date)

    # CPV Codes - unique
    cpv_codes_set = set()
    main_cpv_nodes = root.xpath(".//cac:MainCommodityClassification/cbc:ItemClassificationCode", namespaces=ns)
    for node in main_cpv_nodes:
        if node.text:
            cpv_codes_set.add(node.text.strip())
    add_cpv_nodes = root.xpath(".//cac:AdditionalCommodityClassification/cbc:ItemClassificationCode", namespaces=ns)
    for node in add_cpv_nodes:
        if node.text:
            cpv_codes_set.add(node.text.strip())
    out["CPV Codes"] = ", ".join(sorted(cpv_codes_set))

    return out

def main_scraper(cpv_codes, date_start, date_end, buyer_country, output_excel):
    temp_json = tempfile.mktemp(suffix=".json")
    fetch_all_notices_to_json(cpv_codes, date_start, date_end, buyer_country, temp_json)
    with open(temp_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    notices = data.get("results") or data.get("items") or data.get("notices") or []
    s = requests.Session()
    rows = []
    for n in notices:
        pubno = n.get("publication-number")
        if not pubno:
            continue
        try:
            xml_bytes = fetch_notice_xml(s, pubno, n)
            fields = parse_xml_fields(xml_bytes)
            fields["publication-number"] = pubno
            fields.setdefault("Ted-Link", f"https://ted.europa.eu/en/notice/-/detail/{pubno}")
            rows.append(fields)
        except Exception as e:
            print(f"ERR {pubno}: {e}")
        time.sleep(0.25)
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "publication-number","Beschaffer","Projektbezeichnung","Ort/Region",
        "Vergabeplattform","Ted-Link","Projektstart","Projektende",
        "Geforderte Unternehmensreferenzen","Geforderte Kriterien CVs",
        "Projektvolumen", "Frist Abgabedatum", "Veröffentlichung Datum", "CPV Codes"
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h,"") for h in headers])
    last_row = len(rows) + 1
    last_col = len(headers)
    table_range = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="Teddata", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    wb.save(output_excel)
    os.remove(temp_json)

st.set_page_config(page_title="TED Scraper", layout="centered")
st.markdown("# 📄 TED EU Notice Scraper")
st.markdown("Download TED procurement notices to Excel (data is exported as a table for Power Automate).")
with st.expander("ℹ️ How this works / Instructions", expanded=False):
    st.write("""
    1. Enter your filters (CPV, date range, country, filename).
    2. Click **Run Scraper**. The script downloads notices and attachments, saves an Excel file.
    3. Use the download button to save the Excel file wherever you want!
    4. The exported file now contains an Excel table named 'TEDData', ready for Power Automate!
    """)
c1, c2 = st.columns(2)
with c1:
    cpv_codes = st.text_input("🔎 CPV Codes (space separated)", "71541000 71500000 71240000 79421000 71000000 71248000 71312000 71700000 71300000 71520000 71250000 90712000 71313000")
with c2:
    buyer_country = st.text_input("🌍 Buyer Country (ISO Alpha-3)", "DEU")
today = date.today()
date_col1, date_col2 = st.columns(2)
with date_col1:
    start_date_obj = st.date_input("📆 Start Publication Date", value=date(today.year, today.month, today.day))
with date_col2:
    end_date_obj = st.date_input("📆 End Publication Date", value=date(today.year, today.month, today.day))
date_start = start_date_obj.strftime("%Y%m%d")
date_end   = end_date_obj.strftime("%Y%m%d")
output_excel = st.text_input("💾 Output Excel filename", f"ted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
run = st.button("▶️ Run Scraper")
if run:
    st.info("Scraping... Please wait (can take a few minutes).")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_excel:
        try:
            main_scraper(cpv_codes, date_start, date_end, buyer_country, temp_excel.name)
            st.success("Done! Download your Excel file below.")
            with open(temp_excel.name, "rb") as f:
                st.download_button(
                    label="⬇️ Download Excel",
                    data=f.read(),
                    file_name=output_excel,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Error: {e}")
        finally:
            temp_excel.close()
            os.remove(temp_excel.name)
